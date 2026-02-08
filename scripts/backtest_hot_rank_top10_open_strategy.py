#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
人气榜TOP10开盘买入策略回测引擎

策略逻辑：
1. 前一交易日人气榜前10名
2. 当日开盘价买入（一字涨停跳过）
3. 次日检查：不涨停或人气跌破50则卖出，否则继续持有

使用示例：
    # 使用配置文件
    python scripts/backtest_hot_rank_top10_open_strategy.py \\
        --config config/strategies/hot_rank_top10_open.yaml
    
    # CLI覆盖参数
    python scripts/backtest_hot_rank_top10_open_strategy.py \\
        --config config/strategies/hot_rank_top10_open.yaml \\
        --param.cash_splits=3 \\
        --param.rank_threshold=40
"""

import argparse
import hashlib
import json
import logging
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
import yaml

# 添加项目根目录到路径
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def deep_merge(base: dict, override: dict) -> dict:
    """深度合并字典"""
    result = base.copy()
    for key, value in override.items():
        if key in result and isinstance(result[key], dict) and isinstance(value, dict):
            result[key] = deep_merge(result[key], value)
        else:
            result[key] = value
    return result


def load_strategy_config(config_path: str) -> dict:
    """
    加载策略配置（支持继承）
    
    Args:
        config_path: 策略配置文件路径
        
    Returns:
        合并后的配置字典
    """
    config_path = Path(config_path)
    with open(config_path, 'r', encoding='utf-8') as f:
        config = yaml.safe_load(f)
    
    # 处理继承
    if 'extends' in config:
        base_path = config_path.parent / config['extends']
        with open(base_path, 'r', encoding='utf-8') as f:
            base_config = yaml.safe_load(f)
        config = deep_merge(base_config, config)
        del config['extends']
    
    return config


def apply_cli_overrides(config: dict, overrides: dict) -> dict:
    """
    应用CLI参数覆盖
    
    Args:
        config: 原配置
        overrides: CLI覆盖参数（如 {'param.cash_splits': 3}）
        
    Returns:
        覆盖后的配置
    """
    for key, value in overrides.items():
        if key.startswith('param.'):
            param_name = key[6:]  # 去掉 'param.' 前缀
            if 'params' not in config:
                config['params'] = {}
            config['params'][param_name] = value
            logger.info(f"CLI override: {param_name} = {value}")
    return config


class Trade:
    """交易记录"""
    
    def __init__(self, code: str, entry_date: str, name: str = None):
        self.code = code
        self.name = name
        self.entry_date = entry_date
        self.rank_t_minus_2 = None  # T-2日人气排名
        self.rank_t1 = None      # T-1日人气排名（买入决策依据）
        self.rank_t = None       # T日（买入日）人气排名
        self.rank_t_plus_1 = None  # T+1日人气排名
        self.open_T = None
        self.high_T = None
        self.low_T = None
        self.close_T = None
        self.is_limit_up_open = False  # 是否一字涨停
        self.open_change_pct = None  # 开盘涨跌幅
        
        self.buy_price = None
        self.buy_exec = None
        self.buy_shares = 0
        self.buy_cost = 0
        self.cash_after_buy = 0
        
        self.exit_date = None
        self.exit_reason = None
        self.hold_days = 0
        self.rank_exit = None  # 卖出日人气排名
        self.close_exit = None
        self.is_limit_up_exit = False
        
        self.sell_price = None
        self.sell_exec = None
        self.sell_proceed = 0
        self.cash_after_sell = 0
        
        self.gross_pnl = 0
        self.net_pnl = 0
        self.net_pnl_pct = 0
    
    def to_dict(self) -> dict:
        """转换为字典"""
        return {
            'code': self.code,
            'name': self.name,
            'entry_date': self.entry_date,
            'rank_t_minus_2': self.rank_t_minus_2,
            'rank_t1': self.rank_t1,
            'rank_t': self.rank_t,
            'rank_t_plus_1': self.rank_t_plus_1,
            'open_T': self.open_T,
            'high_T': self.high_T,
            'low_T': self.low_T,
            'close_T': self.close_T,
            'is_limit_up_open': self.is_limit_up_open,
            'open_change_pct': self.open_change_pct,
            'buy_price': self.buy_price,
            'buy_exec': self.buy_exec,
            'buy_shares': self.buy_shares,
            'buy_cost': self.buy_cost,
            'cash_after_buy': self.cash_after_buy,
            'exit_date': self.exit_date,
            'exit_reason': self.exit_reason,
            'hold_days': self.hold_days,
            'rank_exit': self.rank_exit,
            'close_exit': self.close_exit,
            'is_limit_up_exit': self.is_limit_up_exit,
            'sell_price': self.sell_price,
            'sell_exec': self.sell_exec,
            'sell_proceed': self.sell_proceed,
            'cash_after_sell': self.cash_after_sell,
            'gross_pnl': int(round(self.gross_pnl)),
            'net_pnl': self.net_pnl,
            'net_pnl_pct': self.net_pnl_pct
        }


class Position:
    """持仓"""
    
    def __init__(self, trade: Trade):
        self.trade = trade
        self.entry_date = trade.entry_date
        self.code = trade.code
        self.shares = trade.buy_shares
        self.cost_basis = trade.buy_exec
        self.days_held = 0


class BacktestEngine:
    """回测引擎（TOP10开盘买入策略）"""
    
    def __init__(self, config: dict):
        """初始化回测引擎"""
        self.config = config
        self.strategy = config['strategy']
        self.params = config['params']
        self.backtest_config = config['backtest']
        
        # 策略参数
        self.hot_top_n = self.params['hot_top_n']
        self.cash_splits = self.params.get('cash_splits', 3)
        self.per_trade_cash_frac = 1.0 / self.cash_splits
        self.max_positions = self.params.get('max_positions', 3)
        self.rank_threshold = self.params.get('rank_threshold', 50)
        self.max_hold_days = self.params.get('max_hold_days', 30)
        
        # 回测参数
        self.init_cash = self.backtest_config['init_cash']
        self.fee_buy = self.backtest_config['fee_buy']
        self.fee_sell = self.backtest_config['fee_sell']
        self.stamp_tax = self.backtest_config['stamp_tax_sell']
        self.slippage_bps = self.backtest_config['slippage_bps']
        self.min_commission = self.backtest_config['min_commission']
        self.min_lot_size = self.backtest_config['min_lot_size']
        
        # 状态
        self.cash = self.init_cash
        self.positions: Dict[str, Position] = {}
        self.trades: List[Trade] = []
        self.daily_portfolio = []
        self.pending_buy: Dict[str, pd.Series] = {}  # 待买入信号：code -> row_data
        
        # 统计
        self.stats = defaultdict(int)
        
        # 配置日志
        self._setup_logging()
        
        logger.info(f"策略初始化: {self.strategy['name']} v{self.strategy['version']}")
        logger.info(f"参数: hot_top_n={self.hot_top_n}, cash_splits={self.cash_splits}, "
                   f"rank_threshold={self.rank_threshold}")
        logger.info(f"初始资金: {self.init_cash:,.0f}, 分{self.cash_splits}份, "
                   f"每份{self.init_cash * self.per_trade_cash_frac:,.0f}")
    
    def _setup_logging(self):
        """设置日志"""
        # 创建日志目录
        log_dir = PROJECT_ROOT / 'data' / 'backtest' / 'logs'
        log_dir.mkdir(parents=True, exist_ok=True)
        
        # JSON日志文件
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        log_file = log_dir / f"trades_{self.strategy['name']}_{timestamp}.log"
        
        # 添加文件处理器
        self.json_logger = logging.getLogger('trades')
        self.json_logger.setLevel(logging.INFO)
        fh = logging.FileHandler(log_file, encoding='utf-8')
        fh.setFormatter(logging.Formatter('%(message)s'))
        self.json_logger.addHandler(fh)
        
        logger.info(f"交易日志: {log_file}")
    
    def log_trade_event(self, event: str, **kwargs):
        """记录交易事件（JSON格式）"""
        # 转换Timestamp到字符串和numpy类型
        for key, value in kwargs.items():
            if hasattr(value, 'isoformat'):
                kwargs[key] = value.isoformat()
            elif hasattr(value, 'item'):  # numpy types
                kwargs[key] = value.item()
            elif pd.isna(value):  # pandas NA/NaT
                kwargs[key] = None
        
        log_entry = {
            'timestamp': datetime.now().isoformat(),
            'event': event,
            **kwargs
        }
        self.json_logger.info(json.dumps(log_entry, ensure_ascii=False))
    
    def load_features(self, features_path: str) -> pd.DataFrame:
        """加载特征数据"""
        logger.info(f"加载特征数据: {features_path}")
        df = pd.read_parquet(features_path)
        df['date'] = pd.to_datetime(df['date'])
        logger.info(f"加载完成: {len(df):,}行, {df['code'].nunique()}只股票")
        return df
    
    def is_limit_up_board(self, row: pd.Series, prev_close: float) -> bool:
        """
        判断是否一字涨停（无法买入）
        
        一字涨停特征：
        1. 开盘价 = 最高价 = 最低价（一字板）
        2. 价格达到涨停价
        """
        # 检查是否一字板
        is_flat = (row['open'] == row['high'] == row['low'])
        
        # 计算涨停价（根据板块不同）
        code = row['code']
        if code.startswith('688') or code.startswith('689'):  # 科创板
            limit_pct = 0.20
        elif code.startswith('30'):  # 创业板
            limit_pct = 0.20
        elif code.startswith('8') or code.startswith('4'):  # 北交所
            limit_pct = 0.30
        else:  # 主板
            limit_pct = 0.10
        
        limit_up_price = round(prev_close * (1 + limit_pct), 2)
        is_at_limit = row['open'] >= limit_up_price - 0.01  # 允许0.01误差
        
        return is_flat and is_at_limit
    
    def filter_universe(self, df_today: pd.DataFrame, df_prev: pd.DataFrame) -> pd.DataFrame:
        """
        筛选T日人气前10的股票（候选池）
        
        逻辑：
        1. 从T日（df_today）筛选人气前10的股票代码
        2. 过滤ST股票和不可交易股票
        3. 返回候选池（在T+1日执行时再计算涨跌幅并选前3）
        
        Args:
            df_today: T日（今日）特征数据
            df_prev: T-1日（昨日）特征数据（保留参数兼容性）
            
        Returns:
            人气前10的候选股票列表
        """
        # 1. 从T日筛选人气前20的可交易股票
        df_today_hot = df_today[df_today['is_tradable']].copy()
        df_today_hot = df_today_hot[df_today_hot['hot_rank'] <= 20]
        
        if len(df_today_hot) == 0:
            return pd.DataFrame()
        
        self.stats['signal_hot_rank'] += len(df_today_hot)
        
        # 2. 过滤ST股票
        if self.backtest_config.get('filter_st', True):
            before = len(df_today_hot)
            df_today_hot = df_today_hot[~df_today_hot['is_st']]
            self.stats['filter_st'] += (before - len(df_today_hot))
        
        if len(df_today_hot) == 0:
            return pd.DataFrame()
        
        # 3. 按人气排名排序
        df_today_hot = df_today_hot.sort_values('hot_rank', ascending=True).reset_index(drop=True)
        
        return df_today_hot
    
    def execute_buy(self, date: pd.Timestamp, row_today: pd.Series, row_prev: pd.Series, row_prev2: pd.Series = None) -> Optional[Trade]:
        """
        执行买入（开盘价买入）
        
        Args:
            date: 交易日期
            row_today: 今日行情
            row_prev: 昨日行情（T-1）
            row_prev2: 前天行情（T-2），可选
            
        Returns:
            交易记录或None
        """
        code = row_today['code']
        
        # 检查是否一字涨停（如果有昨日收盘价）
        if 'close' in row_prev.index and pd.notna(row_prev['close']):
            if self.is_limit_up_board(row_today, row_prev['close']):
                self.stats['skip_limit_up_open'] += 1
                self.log_trade_event('SKIP_BUY', date=date, code=code, 
                                   reason='limit_up_open', open=row_today['open'])
                return None
        
        # 计算买入价格（开盘价）
        buy_price = row_today['open']
        buy_exec = buy_price * (1 + self.slippage_bps / 10000)
        
        # 计算名义资金（每份）
        nominal_cash = self.init_cash * self.per_trade_cash_frac
        
        # 计算股数（向下取整到100股）
        shares = int(nominal_cash / buy_exec / self.min_lot_size) * self.min_lot_size
        
        if shares == 0:
            self.stats['skip_lot_size'] += 1
            self.log_trade_event('SKIP_BUY', date=date, code=code, 
                               reason='lot_size_insufficient', nominal_cash=nominal_cash)
            return None
        
        # 计算费用
        commission = max(shares * buy_exec * self.fee_buy, self.min_commission)
        total_cost = shares * buy_exec + commission
        
        # 检查现金
        if total_cost > self.cash:
            self.stats['skip_cash'] += 1
            self.log_trade_event('SKIP_BUY', date=date, code=code,
                               reason='insufficient_cash', required=total_cost, available=self.cash)
            return None
        
        # 扣除现金
        self.cash -= total_cost
        
        # 创建交易记录
        trade = Trade(code, date, name=row_today.get('name', ''))
        trade.rank_t_minus_2 = row_prev2.get('hot_rank', None) if row_prev2 is not None and pd.notna(row_prev2.get('hot_rank')) else None
        trade.rank_t1 = row_prev.get('hot_rank', None) if pd.notna(row_prev.get('hot_rank')) else None
        trade.rank_t = row_today.get('hot_rank', None) if pd.notna(row_today.get('hot_rank')) else None
        trade.rank_t_plus_1 = None  # 将在卖出时填充
        trade.open_T = row_today['open']
        trade.high_T = row_today['high']
        trade.low_T = row_today['low']
        trade.close_T = row_today['close']
        trade.is_limit_up_open = False
        # 计算T+1日开盘涨跌幅（T+1日开盘相对T日收盘）
        if 'close' in row_prev.index and pd.notna(row_prev['close']) and row_prev['close'] > 0:
            trade.open_change_pct = (row_today['open'] - row_prev['close']) / row_prev['close']
        
        trade.buy_price = buy_price
        trade.buy_exec = buy_exec
        trade.buy_shares = shares
        trade.buy_cost = total_cost
        trade.cash_after_buy = self.cash
        
        # 添加持仓
        self.positions[code] = Position(trade)
        
        # 记录日志
        self.log_trade_event('BUY', date=date, code=code,
                           rank_t1=trade.rank_t1,
                           buy_price=buy_price,
                           buy_exec=buy_exec,
                           shares=shares,
                           commission=commission,
                           total_cost=total_cost,
                           cash_after=self.cash,
                           reason='open_price')
        
        self.stats['buy_success'] += 1
        logger.info(f"买入: {date} {code} @{buy_exec:.2f} x{shares}股 成本{total_cost:.2f} 余额{self.cash:.2f}")
        
        return trade
    
    def check_exit_signal(self, position: Position, row_today: pd.Series) -> Tuple[bool, str, bool]:
        """
        检查卖出信号
        
        卖出条件（满足任一）：
        1. 未涨停 + 人气跌出前50 -> 开盘卖出
        2. 未涨停 -> 收盘卖出
        3. 达到最大持仓天数 -> 收盘卖出
        
        Args:
            position: 持仓
            row_today: 今日行情
            
        Returns:
            (是否卖出, 卖出原因, 是否开盘卖出)
        """
        # 1. 最大持仓天数
        if position.days_held >= self.max_hold_days:
            return True, 'max_hold_days', False  # 收盘卖出
        
        # 2. 检查涨停
        code = row_today['code']
        if code.startswith('688') or code.startswith('689'):  # 科创板
            limit_pct = 0.20
        elif code.startswith('30'):  # 创业板
            limit_pct = 0.20
        elif code.startswith('8') or code.startswith('4'):  # 北交所
            limit_pct = 0.30
        else:  # 主板
            limit_pct = 0.10
        
        limit_up_price = round(row_today['close_prev'] * (1 + limit_pct), 2)
        is_limit_up = row_today['close'] >= limit_up_price - 0.01
        
        # 3. 检查人气排名
        rank = row_today.get('hot_rank', 999)
        rank_out_top50 = (not pd.isna(rank) and rank > 50)  # 跌出前50
        
        # 判断卖出
        if not is_limit_up and rank_out_top50:
            # 未涨停且跌出前50 -> 开盘卖出
            return True, 'not_limit_up_and_rank_drop', True
        elif not is_limit_up:
            # 未涨停但还在前50 -> 收盘卖出
            return True, 'not_limit_up', False
        elif rank_out_top50:
            # 涨停但跌出前50 -> 不卖（持有涨停股）
            return False, 'hold_limit_up_rank_drop', False
        else:
            # 涨停且排名OK -> 持有
            return False, 'hold_limit_up_rank_ok', False
    
    def execute_sell(self, date: str, position: Position, row_today: pd.Series, reason: str, sell_at_open: bool = False):
        """
        执行卖出
        
        Args:
            date: 交易日期
            position: 持仓
            row_today: 今日行情
            reason: 卖出原因
            sell_at_open: 是否开盘价卖出（True=开盘，False=收盘）
        """
        code = position.code
        trade = position.trade
        
        # 计算卖出价格
        sell_price = row_today['open'] if sell_at_open else row_today['close']
        sell_exec = sell_price * (1 - self.slippage_bps / 10000)
        
        # 计算费用
        commission = max(position.shares * sell_exec * self.fee_sell, self.min_commission)
        stamp = position.shares * sell_exec * self.stamp_tax
        total_fee = commission + stamp
        
        # 计算收益
        sell_proceed = position.shares * sell_exec - total_fee
        self.cash += sell_proceed
        
        # 检查涨停
        code_check = row_today['code']
        if code_check.startswith('688') or code_check.startswith('689'):
            limit_pct = 0.20
        elif code_check.startswith('30'):
            limit_pct = 0.20
        elif code_check.startswith('8') or code_check.startswith('4'):
            limit_pct = 0.30
        else:
            limit_pct = 0.10
        limit_up_price = round(row_today['close_prev'] * (1 + limit_pct), 2)
        is_limit_up = row_today['close'] >= limit_up_price - 0.01
        
        trade.exit_date = date
        trade.exit_reason = reason
        trade.hold_days = position.days_held + 1
        trade.rank_exit = row_today.get('hot_rank', None)
        # 填充T+1日人气排名（即卖出日的人气排名）
        if trade.rank_t_plus_1 is None:
            trade.rank_t_plus_1 = trade.rank_exit
        trade.close_exit = row_today['close']
        trade.is_limit_up_exit = is_limit_up
        trade.sell_price = sell_price
        trade.sell_exec = sell_exec
        trade.sell_proceed = sell_proceed
        trade.cash_after_sell = self.cash
        trade.gross_pnl = sell_proceed - trade.buy_cost
        trade.net_pnl = trade.gross_pnl
        trade.net_pnl_pct = trade.net_pnl / trade.buy_cost
        
        # 记录日志
        self.log_trade_event('SELL', date=date, code=code,
                           exit_reason=reason,
                           sell_price=sell_price,
                           sell_exec=sell_exec,
                           shares=position.shares,
                           commission=commission,
                           stamp_tax=stamp,
                           sell_proceed=sell_proceed,
                           pnl=trade.net_pnl,
                           pnl_pct=trade.net_pnl_pct,
                           hold_days=trade.hold_days,
                           cash_after=self.cash)
        
        self.stats['sell_success'] += 1
        logger.info(f"卖出: {date} {code} @{sell_exec:.2f} x{position.shares}股 "
                   f"收益{trade.net_pnl:.2f}({trade.net_pnl_pct:.2%}) {reason}")
        
        # 移除持仓
        del self.positions[code]
        
        # 添加到已完成交易
        self.trades.append(trade)
    
    def run(self, features_df: pd.DataFrame):
        """
        运行回测
        
        Args:
            features_df: 特征数据
        """
        logger.info("="*80)
        logger.info("开始回测（TOP10开盘买入策略）")
        logger.info("="*80)
        
        # 按日期分组，从2025-01-04开始（避免T-1数据为空）
        all_dates = sorted(features_df['date'].unique())
        start_date = pd.Timestamp('2025-01-04')
        dates = [d for d in all_dates if d >= start_date]
        logger.info(f"回测期间: {dates[0]} 至 {dates[-1]}, 共{len(dates)}个交易日")
        
        for i, date in enumerate(dates):
            # 获取前一个和前两个交易日
            date_idx = all_dates.index(date)
            if date_idx == 0:
                continue  # 第一天没有T-1数据
            prev_date = all_dates[date_idx - 1]
            prev2_date = all_dates[date_idx - 2] if date_idx >= 2 else None
            
            df_today = features_df[features_df['date'] == date]
            df_prev = features_df[features_df['date'] == prev_date]
            df_prev2 = features_df[features_df['date'] == prev2_date] if prev2_date else pd.DataFrame()
            
            logger.info(f"\n--- {date} ---")
            
            # 1. 检查卖出信号（先卖后买）
            positions_to_sell = []
            for code, position in list(self.positions.items()):
                row_today = df_today[df_today['code'] == code]
                if row_today.empty or not row_today.iloc[0]['is_tradable']:
                    # 停牌，继续持有
                    self.log_trade_event('HOLD', date=date, code=code, reason='suspended')
                    position.days_held += 1
                    continue
                
                row_today = row_today.iloc[0]
                should_sell, reason, sell_at_open = self.check_exit_signal(position, row_today)
                
                if should_sell:
                    positions_to_sell.append((position, row_today, reason, sell_at_open))
                elif reason.startswith('hold'):
                    rank = row_today.get('hot_rank', 'N/A')
                    self.log_trade_event('HOLD', date=date, code=code, reason=reason,
                                       close=row_today['close'], rank=rank)
                position.days_held += 1
            
            # 执行卖出
            for position, row_today, reason, sell_at_open in positions_to_sell:
                self.execute_sell(date, position, row_today, reason, sell_at_open)
            
            # 2. 处理昨日pending_buy信号（T+1日执行买入）
            if self.pending_buy:
                # 收集所有候选股票在T+1日的数据
                candidate_codes = list(self.pending_buy.keys())
                candidates = []
                
                for code in candidate_codes:
                    row_today = df_today[df_today['code'] == code]
                    if row_today.empty:
                        logger.warning(f"{code} 昨日产生买入信号，但今日无数据")
                        continue
                    
                    row_today = row_today.iloc[0]
                    
                    # 检查是否可交易
                    if not row_today['is_tradable']:
                        logger.info(f"{code} 昨日产生买入信号，但今日停牌，跳过")
                        continue
                    
                    # 检查是否已持仓
                    if code in self.positions:
                        continue
                    
                    # 计算T+1日开盘涨跌幅（相对T日收盘）
                    row_signal = self.pending_buy[code]
                    if 'close' in row_signal.index and pd.notna(row_signal['close']) and row_signal['close'] > 0:
                        open_change_pct = (row_today['open'] - row_signal['close']) / row_signal['close']
                        candidates.append({
                            'code': code,
                            'row_today': row_today,
                            'row_signal': row_signal,
                            'open_change_pct': open_change_pct
                        })
                
                # 按涨跌幅升序排序（跌幅最大或涨幅最小的在前）
                candidates.sort(key=lambda x: x['open_change_pct'])
                
                # 取前3只（考虑持仓限制）
                n_to_buy = min(3, self.max_positions - len(self.positions))
                selected_candidates = candidates[:n_to_buy]
                
                # 执行买入
                for cand in selected_candidates:
                    code = cand['code']
                    row_today = cand['row_today']
                    row_signal = cand['row_signal']
                    
                    # 获取T-2日数据
                    row_prev2 = df_prev2[df_prev2['code'] == code]
                    row_prev2 = row_prev2.iloc[0] if len(row_prev2) > 0 else None
                    
                    trade = self.execute_buy(date, row_today, row_signal, row_prev2)
                    if trade is None:
                        if self.stats['skip_cash'] > 0:
                            logger.info(f"现金不足，暂时无法买入{code}")
                
                # 清空pending_buy队列
                self.pending_buy.clear()
            
            # 3. 检查新买入信号（T日发现，T+1日执行）
            if len(self.positions) < self.max_positions:
                universe = self.filter_universe(df_today, df_prev)
                logger.info(f"选股池: {len(universe)}只（首次进入前{self.hot_top_n}）")
                
                # filter_universe已经按hot_rank升序排序（1→20，优先买入排名靠前的）
                for _, row_today in universe.iterrows():
                    code = row_today['code']
                    
                    # 检查是否已持仓或已在pending队列
                    if code in self.positions or code in self.pending_buy:
                        continue
                    
                    # 将信号加入pending_buy队列，明天执行
                    self.pending_buy[code] = row_today.copy()
                    logger.info(f"添加买入信号: {code} (T日排名={row_today['hot_rank']}) -> 明日开盘买入")
            
            # 4. 记录每日组合状态
            position_value = sum(
                df_today[df_today['code'] == code].iloc[0]['close'] * pos.shares
                if not df_today[df_today['code'] == code].empty else 0
                for code, pos in self.positions.items()
            )
            nav = self.cash + position_value
            
            self.daily_portfolio.append({
                'date': date,
                'cash': self.cash,
                'position_value': position_value,
                'nav': nav,
                'n_positions': len(self.positions)
            })
            
            logger.info(f"持仓: {len(self.positions)}只, 现金: {self.cash:.2f}, "
                       f"市值: {position_value:.2f}, 净值: {nav:.2f}")
        
        logger.info("="*80)
        logger.info("回测完成")
        logger.info("="*80)
        self.print_stats()
    
    def print_stats(self):
        """打印统计信息"""
        logger.info(f"\n统计信息:")
        logger.info(f"  信号数（人气前{self.hot_top_n}）: {self.stats['signal_hot_rank']}")
        logger.info(f"  过滤-ST股票: {self.stats.get('filter_st', 0)}")
        logger.info(f"  跳过-一字涨停: {self.stats.get('skip_limit_up_open', 0)}")
        logger.info(f"  成交买入: {self.stats['buy_success']}")
        logger.info(f"  成交卖出: {self.stats['sell_success']}")
        logger.info(f"  跳过-资金不足: {self.stats['skip_cash']}")
        logger.info(f"  跳过-不足一手: {self.stats['skip_lot_size']}")
        logger.info(f"  最终现金: {self.cash:.2f}")
        logger.info(f"  最终持仓: {len(self.positions)}只")
    
    def save_formatted_excel(self, df: pd.DataFrame, filepath: Path):
        """
        保存格式化的Excel文件（带颜色和样式）
        
        Args:
            df: 交易数据DataFrame
            filepath: 输出文件路径
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            logger.warning("未安装openpyxl，跳过Excel格式化导出。可使用 pip install openpyxl 安装")
            return
        
        wb = Workbook()
        ws = wb.active
        ws.title = "交易明细"
        
        # 写入数据
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                # 处理net_pnl_pct列：转换为百分比
                if r_idx > 1 and 'net_pnl_pct' in df.columns and c_idx == df.columns.get_loc('net_pnl_pct') + 1:
                    if isinstance(value, (int, float)):
                        value = f"{value * 100:.2f}%"
                
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # 表头样式
                if r_idx == 1:
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF", size=11)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    # 交替行颜色
                    if r_idx % 2 == 0:
                        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                    
                    # 根据盈亏上色
                    if 'net_pnl' in df.columns and c_idx == df.columns.get_loc('net_pnl') + 1:
                        if isinstance(value, (int, float)) and value != 0:
                            if value > 0:
                                cell.font = Font(color="00AA00", bold=True)  # 绿色
                            elif value < 0:
                                cell.font = Font(color="FF0000", bold=True)  # 红色
                    
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # 边框
                thin_border = Border(
                    left=Side(style='thin', color='CCCCCC'),
                    right=Side(style='thin', color='CCCCCC'),
                    top=Side(style='thin', color='CCCCCC'),
                    bottom=Side(style='thin', color='CCCCCC')
                )
                cell.border = thin_border
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 冻结首行
        ws.freeze_panes = 'A2'
        
        wb.save(filepath)
    
    def save_results(self, output_dir: str):
        """保存回测结果"""
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        config_hash = hashlib.md5(json.dumps(self.params, sort_keys=True).encode()).hexdigest()[:8]
        prefix = f"{self.strategy['name']}_v{self.strategy['version']}_{config_hash}_{timestamp}"
        
        # 保存交易明细
        if self.trades:
            trades_df = pd.DataFrame([t.to_dict() for t in self.trades])
            
            # 调整列顺序：将盈亏列移到前面，买卖价格和日期挨着
            cols = list(trades_df.columns)
            # 定义前置列（关键字段）
            priority_cols = ['code', 'name', 'entry_date', 'exit_date', 'gross_pnl', 'net_pnl', 'net_pnl_pct', 'open_change_pct',
                           'buy_price', 'buy_exec', 'sell_price', 'sell_exec']
            # 移除已存在的前置列
            remaining_cols = [c for c in cols if c not in priority_cols]
            # 重新排序：前置列 + 其他列
            new_order = [c for c in priority_cols if c in cols] + remaining_cols
            trades_df = trades_df[new_order]
            
            # 按entry_date降序排列（最新的在前）
            trades_df = trades_df.sort_values('entry_date', ascending=False)
            
            # 格式化数字：保留2位小数
            numeric_cols = trades_df.select_dtypes(include=['float64', 'float32']).columns
            for col in numeric_cols:
                if col != 'net_pnl_pct':  # net_pnl_pct单独处理
                    trades_df[col] = trades_df[col].round(2)
            
            trades_file = output_dir / 'trades' / f"{prefix}_trades.parquet"
            trades_file.parent.mkdir(exist_ok=True)
            trades_df.to_parquet(trades_file, index=False)
            logger.info(f"交易明细已保存: {trades_file}")
            
            # 同时保存CSV
            csv_file = output_dir / 'trades' / f"{prefix}_trades.csv"
            trades_df.to_csv(csv_file, index=False, encoding='utf-8-sig')
            logger.info(f"交易明细CSV已保存: {csv_file}")
            
            # 保存格式化的Excel文件
            xlsx_file = output_dir / 'trades' / f"{prefix}_trades.xlsx"
            self.save_formatted_excel(trades_df, xlsx_file)
            logger.info(f"格式化Excel已保存: {xlsx_file}")
        
        # 保存组合净值
        if self.daily_portfolio:
            portfolio_df = pd.DataFrame(self.daily_portfolio)
            portfolio_file = output_dir / 'portfolio' / f"{prefix}_portfolio.parquet"
            portfolio_file.parent.mkdir(exist_ok=True)
            portfolio_df.to_parquet(portfolio_file, index=False)
            logger.info(f"组合净值已保存: {portfolio_file}")


def main():
    parser = argparse.ArgumentParser(
        description='人气榜TOP10开盘买入策略回测引擎',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__
    )
    
    parser.add_argument(
        '--config',
        required=True,
        help='策略配置文件路径'
    )
    parser.add_argument(
        '--features',
        default='data/processed/features/daily_features_v1.parquet',
        help='特征数据路径'
    )
    parser.add_argument(
        '--output',
        default='data/backtest',
        help='输出目录'
    )
    
    # CLI参数覆盖
    parser.add_argument('--param.cash_splits', type=int, dest='param_cash_splits')
    parser.add_argument('--param.rank_threshold', type=int, dest='param_rank_threshold')
    
    args = parser.parse_args()
    
    # 加载配置
    config = load_strategy_config(args.config)
    
    # 应用CLI覆盖
    cli_overrides = {}
    if args.param_cash_splits:
        cli_overrides['param.cash_splits'] = args.param_cash_splits
    if args.param_rank_threshold:
        cli_overrides['param.rank_threshold'] = args.param_rank_threshold
    
    if cli_overrides:
        config = apply_cli_overrides(config, cli_overrides)
    
    # 初始化回测引擎
    engine = BacktestEngine(config)
    
    # 加载特征数据
    features_df = engine.load_features(args.features)
    
    # 运行回测
    engine.run(features_df)
    
    # 保存结果
    engine.save_results(args.output)


if __name__ == '__main__':
    main()
